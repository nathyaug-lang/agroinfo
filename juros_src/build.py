"""Gera index.html e juros.xlsx com decisões do FOMC e do Copom desde 2003.

Saída:
- index.html  — relatório navegável (tabelas + 3 gráficos SVG)
- juros.xlsx  — pasta com 4 abas (FOMC, Selic, Diferencial, Próximas reuniões)

Estilo visual: Innovagro Research & Market Intelligence (ver ESTILO_VISUAL.md).
"""

from __future__ import annotations

import base64
import datetime as dt
import json
import time
import urllib.error
import urllib.request
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from dados import FOMC, PROXIMAS_COPOM, PROXIMAS_FOMC
from fetch_dados import (
    balanca_comercial_br,
    balanca_comercial_us,
    divida_pib_br,
    dollar_index,
    fluxo_capital_estrangeiro,
    focus_expectativas,
    usd_brl,
)

ROOT = Path(__file__).resolve().parent
HOJE = dt.date.today()  # data de execução; atualiza sozinha a cada `python build.py`


# ---------------------------------------------------------------------------
# helpers para séries macro
# ---------------------------------------------------------------------------

def _to_dates(serie: list[tuple[str, float]]) -> list[tuple[dt.date, float]]:
    return [(dt.date.fromisoformat(d), v) for d, v in serie]


def _amostra_mensal(serie: list[tuple[dt.date, float]]) -> list[tuple[dt.date, float]]:
    """Mantém o último valor de cada mês — útil para reduzir séries diárias."""
    bucket: dict[tuple[int, int], tuple[dt.date, float]] = {}
    for d, v in serie:
        bucket[(d.year, d.month)] = (d, v)
    return sorted(bucket.values(), key=lambda r: r[0])


def _balanca_acum12m(serie: list[tuple[dt.date, float]]) -> list[tuple[dt.date, float]]:
    """Soma móvel de 12 meses — leitura clássica de balança comercial."""
    if len(serie) < 12:
        return list(serie)
    out = []
    valores = [v for _, v in serie]
    for i in range(11, len(serie)):
        acum = sum(valores[i - 11:i + 1])
        out.append((serie[i][0], acum))
    return out

# ---------------------------------------------------------------------------
# Selic-Meta — fonte: SGS/BCB série 432
# ---------------------------------------------------------------------------

SELIC_HISTORY_CACHE = ROOT / ".selic_history_cache.json"
SELIC_FRESHNESS_HOURS = 12

# Anotações editorais que enriquecem a observação automática (+/- bp)
# para datas com contexto histórico relevante. Usa data ISO de vigência.
ANOTACOES_SELIC = {
    "2008-10-30": "COVID antecede · auge da crise financeira (set/2008)",
    "2009-01-22": "Início do ciclo de cortes pós-crise",
    "2020-03-19": "COVID-19 · corte de emergência",
    "2020-08-06": "Mínima histórica · piso da Selic",
    "2021-03-18": "Início do ciclo de alta pós-pandemia",
    "2024-09-19": "Retomada da alta · pressão fiscal e câmbio",
}


def _fetch_serie(codigo: int, ini: str, fim: str) -> list[dict]:
    url = (
        f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.{codigo}/dados"
        f"?formato=json&dataInicial={ini}&dataFinal={fim}"
    )
    req = urllib.request.Request(
        url, headers={"User-Agent": "Innovagro/1.0", "Accept": "application/json"}
    )
    last_err: Exception | None = None
    for tentativa in range(3):
        try:
            with urllib.request.urlopen(req, timeout=40) as r:
                return json.loads(r.read().decode("utf-8"))
        except (urllib.error.URLError, TimeoutError) as e:
            last_err = e
            time.sleep(2 ** tentativa)
    raise RuntimeError(f"BCB SGS {codigo} {ini}-{fim}: {last_err}") from last_err


def _detect_changes(daily: list[dict]) -> list[tuple[str, float]]:
    """Reduz a série diária a pontos de mudança de valor."""
    out: list[tuple[str, float]] = []
    prev: float | None = None
    for d in daily:
        v = float(d["valor"])
        if v != prev:
            out.append((d["data"], v))
            prev = v
    return out


def _enrich_selic(mudancas: list[tuple[str, float]]) -> list[tuple[str, float, str, str]]:
    """Converte mudanças (dd/mm/yyyy, valor) em tuplas no formato do projeto."""
    pares = []
    valores = [v for _, v in mudancas]
    if not valores:
        return []
    v_max, v_min = max(valores), min(valores)

    for i, (data_br, valor) in enumerate(mudancas):
        d, m, y = data_br.split("/")
        data_iso = f"{y}-{m}-{d}"

        if i == 0:
            obs = "Vigente em 2003-01-01"
        else:
            delta = round((valor - mudancas[i - 1][1]) * 100)  # bps
            if delta > 0:
                obs = f"+{delta} bp"
            elif delta < 0:
                obs = f"{delta} bp"
            else:
                obs = "Mantém"

        # marcadores de extremos absolutos no período
        if valor == v_max:
            obs = f"{obs} · pico (≈{valor:.2f}%)"
        elif valor == v_min:
            obs = f"{obs} · mínima (≈{valor:.2f}%)"

        # anotação editorial sobrescreve/complementa
        anot = ANOTACOES_SELIC.get(data_iso)
        if anot:
            obs = f"{obs} · {anot}"

        pares.append((data_iso, valor, "Reunião regular", obs))
    return pares


def fetch_selic_history() -> list[tuple[str, float, str, str]]:
    """Devolve a série de mudanças da Selic-Meta (BCB SGS 432) desde 2003.

    Cache local em `.selic_history_cache.json`; recarrega se mais antigo
    que `SELIC_FRESHNESS_HOURS`. Em caso de falha de rede usa o cache mesmo
    que velho, e levanta exceção apenas se nem cache existe.
    """
    cache_fresh = False
    if SELIC_HISTORY_CACHE.exists():
        idade_h = (
            time.time() - SELIC_HISTORY_CACHE.stat().st_mtime
        ) / 3600
        cache_fresh = idade_h < SELIC_FRESHNESS_HOURS

    if cache_fresh:
        mudancas = json.loads(SELIC_HISTORY_CACHE.read_text(encoding="utf-8"))
        return _enrich_selic(mudancas)

    try:
        chunks = [
            ("01/01/2003", "31/12/2007"),
            ("01/01/2008", "31/12/2012"),
            ("01/01/2013", "31/12/2017"),
            ("01/01/2018", "31/12/2022"),
            ("01/01/2023", HOJE.strftime("%d/%m/%Y")),
        ]
        diario: list[dict] = []
        for ini, fim in chunks:
            diario.extend(_fetch_serie(432, ini, fim))
        mudancas = _detect_changes(diario)
        SELIC_HISTORY_CACHE.write_text(
            json.dumps(mudancas, ensure_ascii=False, indent=1),
            encoding="utf-8",
        )
        return _enrich_selic(mudancas)
    except Exception as e:
        if SELIC_HISTORY_CACHE.exists():
            print(f"[aviso] BCB SGS indisponível ({e}); usando cache local.")
            mudancas = json.loads(SELIC_HISTORY_CACHE.read_text(encoding="utf-8"))
            return _enrich_selic(mudancas)
        raise


SELIC = fetch_selic_history()

# Logo institucional Innovagro Brasil — usada como marca-d'água nos charts.
LOGO_PATH = (
    ROOT.parent / "SnD_Sistema" / "assets" / "logo_innovagro.jpg"
)


def logo_data_uri() -> str:
    """Logo removida — projeto sem marca. Retorna string vazia.

    Mantida como função (em vez de remover) para não quebrar referências a
    LOGO_URI no restante do código. Sem logo, a marca-d'água dos charts e o
    logo do cabeçalho simplesmente não são renderizados.
    """
    return ""


LOGO_URI = logo_data_uri()


# ---------------------------------------------------------------------------
# Helpers numéricos
# ---------------------------------------------------------------------------

def parse_date(s: str) -> dt.date:
    return dt.date.fromisoformat(s)


def fmt_date_br(s: str) -> str:
    d = parse_date(s)
    return f"{d.day:02d}/{d.month:02d}/{d.year}"


def fmt_pct_br(v: float, casas: int = 2) -> str:
    return f"{v:.{casas}f}".replace(".", ",") + "%"


def serie_mensal(series: list[tuple[str, float, str, str]],
                 inicio: dt.date,
                 fim: dt.date) -> list[tuple[dt.date, float]]:
    """Forward-fill mensal de uma série de decisões para alinhar Fed e Selic."""
    out: list[tuple[dt.date, float]] = []
    pares = [(parse_date(d), v) for d, v, *_ in series]
    pares.sort()
    idx = 0
    atual = pares[0][1]
    cur = dt.date(inicio.year, inicio.month, 1)
    while cur <= fim:
        while idx < len(pares) and pares[idx][0] <= cur:
            atual = pares[idx][1]
            idx += 1
        out.append((cur, atual))
        cur = (dt.date(cur.year + (cur.month // 12), ((cur.month % 12) + 1), 1)
               if cur.month != 12 else dt.date(cur.year + 1, 1, 1))
    return out


# ---------------------------------------------------------------------------
# SVG chart engine — linha temporal com marca-d'água, eixos navy + grid
# ---------------------------------------------------------------------------

def _wm_image(uid: str, chart_w: int, chart_h: int) -> str:
    """Marca-d'água com a logo Innovagro Brasil (jpg embutido em base64).

    A imagem é centralizada, sem rotação, com opacidade 7% — segue o
    padrão definido no ESTILO_VISUAL.md (item 5.4).
    """
    # Logo removida — sem marca-d'água nos charts.
    if not LOGO_URI:
        return ""
    # logo ocupa ~52% da largura do chart, mantendo proporção 3580x733 (~4.88:1)
    img_w = int(chart_w * 0.52)
    img_h = int(img_w * 733 / 3580)
    x = (chart_w - img_w) / 2
    y = (chart_h - img_h) / 2
    return (
        f'<image id="wm{uid}" href="{LOGO_URI}" x="{x:.1f}" y="{y:.1f}" '
        f'width="{img_w}" height="{img_h}" opacity="0.07" '
        f'preserveAspectRatio="xMidYMid meet"/>'
    )


def chart_linha(serie: list[tuple[dt.date, float]],
                titulo: str,
                cor: str,
                uid: str,
                y_min: float | None = None,
                y_max: float | None = None,
                rotular_ano_min: int = 2003,
                eixo_zero: bool = False,
                export: bool = False) -> str:
    """Renderiza um chart de linha em SVG (estilo Innovagro).

    `export=True` gera uma versão com proporção mais quadrada (2:1),
    fontes maiores e maior espaçamento entre rótulos — pensada para a
    cópia do PNG, onde o chart fica em ~900 px de largura final.
    """
    if export:
        W, H = 1100, 540
        PAD_T, PAD_R, PAD_B, PAD_L = 42, 56, 110, 130
        F_AXIS, F_LABEL, F_DATE = 22, 26, 18
        MIN_PX = 195       # ~21% da largura útil — caber extremos importantes
        ANO_STEP = 4   # export: rotula 1 ano a cada 4 (mais limpo)
    else:
        W, H = 1480, 360
        PAD_T, PAD_R, PAD_B, PAD_L = 30, 36, 64, 70
        F_AXIS, F_LABEL, F_DATE = 10, 10.5, 9
        MIN_PX = 78
        ANO_STEP = 2   # tela (não usado — preserva lógica original abaixo)

    xs = [d for d, _ in serie]
    ys = [v for _, v in serie]
    x_min = xs[0].toordinal()
    x_max = xs[-1].toordinal()
    if y_min is None:
        y_min = min(ys)
    if y_max is None:
        y_max = max(ys)
    span = max(y_max - y_min, 0.01)
    pad = span * 0.10
    y_min_p, y_max_p = y_min - pad, y_max + pad
    if eixo_zero and y_min_p > 0:
        y_min_p = min(y_min_p, -0.5)

    plot_w = W - PAD_L - PAD_R
    plot_h = H - PAD_T - PAD_B

    def fx(d: dt.date) -> float:
        return PAD_L + (d.toordinal() - x_min) / max(x_max - x_min, 1) * plot_w

    def fy(v: float) -> float:
        return PAD_T + (1 - (v - y_min_p) / (y_max_p - y_min_p)) * plot_h

    # ---------- eixos / grid ----------
    grid_steps = 6
    grid_lines = []
    y_labels = []
    for i in range(grid_steps + 1):
        v = y_min_p + (y_max_p - y_min_p) * i / grid_steps
        y = fy(v)
        grid_lines.append(
            f'<line x1="{PAD_L}" y1="{y:.1f}" x2="{W-PAD_R}" y2="{y:.1f}" '
            f'stroke="#E4E7EB" stroke-width="1"/>'
        )
        y_labels.append(
            f'<text x="{PAD_L-8}" y="{y+3:.1f}" text-anchor="end" '
            f'font-size="{F_AXIS}" fill="#5B6775">{fmt_pct_br(v, 1)}</text>'
        )

    # zero baseline em destaque
    if y_min_p < 0 < y_max_p:
        y0 = fy(0)
        grid_lines.append(
            f'<line x1="{PAD_L}" y1="{y0:.1f}" x2="{W-PAD_R}" y2="{y0:.1f}" '
            f'stroke="#0A2342" stroke-width="1.2" stroke-dasharray="2,3"/>'
        )

    # ---------- ticks no eixo X (anos) ----------
    x_ticks = []
    ano_ini = xs[0].year
    ano_fim = xs[-1].year
    for ano in range(ano_ini, ano_fim + 1):
        if export:
            # export: 1 ano a cada ANO_STEP a partir do início + extremos
            pular = (ano - ano_ini) % ANO_STEP != 0
        else:
            # tela: comportamento original — anos pares + extremos
            pular = ano % 2 != 0
        if pular and ano not in (ano_ini, ano_fim):
            continue
        d = dt.date(ano, 1, 1)
        if d < xs[0]:
            d = xs[0]
        if d > xs[-1]:
            d = xs[-1]
        x = fx(d)
        x_ticks.append(
            f'<line x1="{x:.1f}" y1="{H-PAD_B}" x2="{x:.1f}" y2="{H-PAD_B+5}" '
            f'stroke="#0A2342" stroke-width="1"/>'
            f'<text x="{x:.1f}" y="{H-PAD_B+(28 if export else 18):.1f}" text-anchor="middle" '
            f'font-size="{F_AXIS}" fill="#5B6775">{ano}</text>'
        )

    # ---------- linha ----------
    pts = [f"{fx(d):.1f},{fy(v):.1f}" for d, v in serie]
    path_d = "M" + " L".join(pts)

    # área de leve preenchimento sob a linha
    bottom_y = fy(y_min_p)
    area = (
        f'<polygon points="{pts[0].split(",")[0]},{bottom_y:.1f} '
        f'{" ".join(pts)} '
        f'{pts[-1].split(",")[0]},{bottom_y:.1f}" '
        f'fill="{cor}" fill-opacity="0.08"/>'
    )

    linha = (
        f'<path d="{path_d}" fill="none" stroke="{cor}" stroke-width="2.2" '
        f'stroke-linejoin="round" stroke-linecap="round"/>'
    )

    # ---------- rótulos: primeiro, último, máximo, mínimo, picos/vales ----------
    rot_idx = set()
    rot_idx.add(0)
    rot_idx.add(len(serie) - 1)
    # max & min absolutos
    i_max = max(range(len(serie)), key=lambda i: serie[i][1])
    i_min = min(range(len(serie)), key=lambda i: serie[i][1])
    rot_idx.add(i_max)
    rot_idx.add(i_min)
    # extremos locais com janela ampla — rotula picos/vales relevantes
    JANELA = max(6, len(serie) // 14)
    for i in range(JANELA, len(serie) - JANELA):
        v = serie[i][1]
        viz = [serie[j][1] for j in range(i - JANELA, i + JANELA + 1) if j != i]
        if v == max(serie[i - JANELA:i + JANELA + 1], key=lambda t: t[1])[1] and v - min(viz) > span * 0.10:
            rot_idx.add(i)
        if v == min(serie[i - JANELA:i + JANELA + 1], key=lambda t: t[1])[1] and max(viz) - v > span * 0.10:
            rot_idx.add(i)

    # de-clutter: descarta rótulos muito próximos no eixo X
    if export:
        # Export: primeiro e último ponto são âncoras obrigatórias (mostram
        # o início da janela e o valor atual). Os demais entram em ordem de
        # relevância (mais extremos primeiro) somente se não colidirem com
        # nenhum já mantido.
        ancoras = [0, len(serie) - 1]
        filtrados: list[int] = list(ancoras)
        outros = sorted(
            rot_idx - set(ancoras),
            key=lambda i: -abs(serie[i][1] - (y_min + y_max) / 2),
        )
        for i in outros:
            xi = fx(serie[i][0])
            if not any(abs(xi - fx(serie[j][0])) < MIN_PX for j in filtrados):
                filtrados.append(i)
        filtrados.sort()
    else:
        # Tela: comportamento original (mais extremo ganha em colisão)
        rot = sorted(rot_idx)
        filtrados = []
        for i in rot:
            if not filtrados or fx(serie[i][0]) - fx(serie[filtrados[-1]][0]) >= MIN_PX:
                filtrados.append(i)
            else:
                j = filtrados[-1]
                if abs(serie[i][1] - (y_min + y_max) / 2) > abs(serie[j][1] - (y_min + y_max) / 2):
                    filtrados[-1] = i

    pontos = []
    rotulos = []
    R_PONTO = 6.0 if export else 3.6
    OFF_ACIMA = -26 if export else -9
    OFF_ABAIXO = 42 if export else 16
    OFF_DATA = 28 if export else 12
    for i in filtrados:
        d, v = serie[i]
        x, y = fx(d), fy(v)
        pontos.append(
            f'<circle cx="{x:.1f}" cy="{y:.1f}" r="{R_PONTO}" fill="#fff" '
            f'stroke="{cor}" stroke-width="2"/>'
        )
        # Posiciona rótulo: respeita curva, mas força "acima" quando o ponto
        # está perto do eixo X (senão o rótulo abaixo colide com os anos)
        # e força "abaixo" quando está colado no topo (colide com o título).
        y_rel = (y - PAD_T) / plot_h          # 0=topo, 1=fundo
        natural_acima = i == 0 or serie[i][1] >= serie[i - 1][1]
        if y_rel > 0.78:
            acima = True
        elif y_rel < 0.22:
            acima = False
        else:
            acima = natural_acima
        ty = y + (OFF_ACIMA if acima else OFF_ABAIXO)
        rotulos.append(
            f'<text x="{x:.1f}" y="{ty:.1f}" text-anchor="middle" '
            f'font-size="{F_LABEL}" font-weight="700" fill="{cor}">'
            f'{fmt_pct_br(v, 2)}</text>'
        )
        rotulos.append(
            f'<text x="{x:.1f}" y="{ty + (-OFF_DATA if acima else OFF_DATA):.1f}" '
            f'text-anchor="middle" font-size="{F_DATE}" fill="#5B6775">{d.strftime("%b/%y")}</text>'
        )

    # ---------- caixa final ----------
    eixo = (
        f'<line x1="{PAD_L}" y1="{H-PAD_B}" x2="{W-PAD_R}" y2="{H-PAD_B}" '
        f'stroke="#0A2342" stroke-width="1.5"/>'
        f'<line x1="{PAD_L}" y1="{PAD_T}" x2="{PAD_L}" y2="{H-PAD_B}" '
        f'stroke="#0A2342" stroke-width="1.5"/>'
    )

    klass = "chart-svg-export" if export else "chart-svg"
    svg = f"""
<svg viewBox=\"0 0 {W} {H}\" xmlns=\"http://www.w3.org/2000/svg\" role=\"img\"
     aria-label=\"{titulo}\" class=\"{klass}\" data-chart-id=\"{uid}\"
     data-w=\"{W}\" data-h=\"{H}\">
  {_wm_image(uid, W, H)}
  {''.join(grid_lines)}
  {area}
  {linha}
  {eixo}
  {''.join(x_ticks)}
  {''.join(y_labels)}
  {''.join(pontos)}
  {''.join(rotulos)}
</svg>
"""
    return svg.strip()


# ---------------------------------------------------------------------------
# Tabelas HTML
# ---------------------------------------------------------------------------

def _direcao(obs: str) -> tuple[str, str]:
    """Devolve (símbolo, classe-pill) com base na observação textual."""
    obs_l = obs.lower()
    if "+" in obs and "bp" in obs_l:
        return "▲", "alta"
    if "-" in obs and "bp" in obs_l:
        return "▼", "baixa"
    if "manté" in obs_l or "mantém" in obs_l or "banda" in obs_l:
        return "■", "neutro"
    return "■", "neutro"


def html_tabela_decisoes(decisoes, titulo_curto: str) -> str:
    rows = []
    for data, taxa, _evento, obs in reversed(decisoes):  # mais recente em cima
        sym, cls = _direcao(obs)
        rows.append(
            f"<tr><td class=\"col-data\">{fmt_date_br(data)}</td>"
            f"<td class=\"col-taxa\"><span class=\"pill {cls}\">{sym} {fmt_pct_br(taxa, 2)}</span></td>"
            f"<td class=\"col-mov\">{obs}</td></tr>"
        )
    return (
        f'<div class="tabela-wrap" data-tabela="{titulo_curto}">'
        '<table>'
        '<thead><tr>'
        '<th class="col-data">Data</th>'
        '<th class="col-taxa">Taxa-meta</th>'
        '<th class="col-mov">Movimento</th>'
        '</tr></thead>'
        f'<tbody>{"".join(rows)}</tbody>'
        '</table></div>'
    )


def chart_card(titulo: str, meta: str, svg: str, svg_export: str,
               filename: str) -> str:
    """Empacota um chart com cabeçalho (título + botão de copiar imagem).

    `svg_export` é uma versão alternativa (proporção mais quadrada e fontes
    maiores) que vai escondida no DOM e é usada apenas pelo botão de copiar.
    """
    return f"""
  <div class=\"chart-card\">
    <div class=\"chart-head\">
      <div>
        <div class=\"ch-title\">{titulo}</div>
        <div class=\"ch-meta\">{meta}</div>
      </div>
      <button type=\"button\" class=\"copy-btn\" data-filename=\"{filename}\"
              aria-label=\"Copiar gráfico como imagem\"
              title=\"Copiar gráfico como imagem\">
        <svg width=\"15\" height=\"15\" viewBox=\"0 0 24 24\" fill=\"none\"
             stroke=\"currentColor\" stroke-width=\"2\"
             stroke-linecap=\"round\" stroke-linejoin=\"round\">
          <rect x=\"9\" y=\"9\" width=\"13\" height=\"13\" rx=\"2\"/>
          <path d=\"M5 15H4a2 2 0 0 1-2-2V4a2 2 0 0 1 2-2h9a2 2 0 0 1 2 2v1\"/>
        </svg>
      </button>
    </div>
    <div class=\"ch-svg\">{svg}</div>
    <div class=\"ch-svg-export-wrap\" hidden aria-hidden=\"true\">{svg_export}</div>
  </div>"""


def html_tabela_proximas() -> str:
    pares = []
    for data, _evt, proj in PROXIMAS_FOMC:
        pares.append((parse_date(data), "FOMC · Federal Reserve", proj))
    for data, _evt, proj in PROXIMAS_COPOM:
        pares.append((parse_date(data), "Copom · Banco Central do Brasil", proj))
    pares.sort(key=lambda p: p[0])

    rows = []
    for d, comite, proj in pares:
        d_str = d.strftime("%d/%m/%Y")
        dias = (d - HOJE).days
        dias_str = f"{dias} d" if dias >= 0 else "—"
        proj_html = proj if proj and proj.strip() else "—"
        rows.append(
            f"<tr><td class=\"col-data\">{d_str}</td>"
            f"<td class=\"col-comite\">{comite}</td>"
            f"<td class=\"col-proj\">{proj_html}</td>"
            f"<td class=\"col-faltam\">{dias_str}</td></tr>"
        )
    return (
        '<div class="tabela-wrap" data-tabela="proximas">'
        '<table>'
        '<thead><tr>'
        '<th class="col-data">Data prevista</th>'
        '<th class="col-comite">Comitê</th>'
        '<th class="col-proj">Projeção</th>'
        '<th class="col-faltam">Faltam</th>'
        '</tr></thead>'
        f'<tbody>{"".join(rows)}</tbody>'
        '</table></div>'
    )


# ---------------------------------------------------------------------------
# Seção MACRO — dados puxados em runtime via fetch_dados
# ---------------------------------------------------------------------------

def _fmt_us_bi(v: float, casas: int = 1) -> str:
    """Formata em bilhões de US$ no estilo BR (- US$ 909,6 bi)."""
    return f"US$ {v / 1e9:,.{casas}f} bi".replace(",", "X").replace(".", ",").replace("X", ".")


def _fmt_us_mi(v: float, casas: int = 1) -> str:
    """Formata em milhões de US$."""
    return f"US$ {v:,.{casas}f} mi".replace(",", "X").replace(".", ",").replace("X", ".")


def _carregar_macro() -> dict:
    """Puxa as 7 séries macro com cache (silenciosamente — se algo falhar,
    a série retorna lista vazia e o card mostra placeholder)."""
    out: dict = {}
    for nome, fn in (
        ("divida_pib", divida_pib_br),
        ("usd_brl_d", usd_brl),
        ("dxy", dollar_index),
        ("balanca_br", balanca_comercial_br),
        ("balanca_us", balanca_comercial_us),
        ("fluxo_cap", fluxo_capital_estrangeiro),
    ):
        try:
            out[nome] = _to_dates(fn())
        except Exception as e:
            print(f"[macro] falha em {nome}: {e}")
            out[nome] = []
    try:
        out["focus"] = focus_expectativas()
    except Exception as e:
        print(f"[macro] falha em focus: {e}")
        out["focus"] = {}
    return out


def _chart_par(serie, titulo, cor, uid, **kw) -> tuple[str, str]:
    """Devolve (svg_tela, svg_export) para uma série."""
    if not serie:
        return ("", "")
    s_tela = chart_linha(serie, titulo, cor, uid, **kw)
    s_exp = chart_linha(serie, titulo, cor, f"{uid}x",
                        export=True, **kw)
    return s_tela, s_exp


def html_secao_macro(macro: dict) -> str:
    blocos: list[str] = []

    # 1. Dívida Bruta / PIB
    s = macro["divida_pib"]
    if s:
        atual = s[-1]; pico = max(s, key=lambda r: r[1]); minima = min(s, key=lambda r: r[1])
        tela, exp = _chart_par(
            s, "Dívida Bruta do Governo Geral (% PIB)", "#0A2342", uid="dbgg",
        )
        blocos.append(chart_card(
            "Dívida Bruta / PIB (DBGG)",
            f"Atual: <strong>{fmt_pct_br(atual[1], 2)}</strong> em {fmt_date_br(atual[0].isoformat())} · "
            f"Pico: <strong>{fmt_pct_br(pico[1], 2)}</strong> em {fmt_date_br(pico[0].isoformat())} · "
            f"Fonte: BCB SGS · série 13762",
            tela, exp, "divida-pib-br",
        ))

    # 2. USD/BRL (amostragem mensal)
    s = _amostra_mensal(macro["usd_brl_d"])
    if s:
        atual = s[-1]; pico = max(s, key=lambda r: r[1]); minima = min(s, key=lambda r: r[1])
        tela, exp = _chart_par(s, "USD/BRL (PTAX venda · mensal)", "#B45309", uid="usdbrl")
        blocos.append(chart_card(
            "Câmbio USD/BRL",
            f"Atual: <strong>R$ {atual[1]:.4f}</strong> em {fmt_date_br(atual[0].isoformat())} · "
            f"Pico: <strong>R$ {pico[1]:.4f}</strong> em {fmt_date_br(pico[0].isoformat())} · "
            f"Mínima: <strong>R$ {minima[1]:.4f}</strong> · "
            f"Fonte: BCB SGS · série 1 (PTAX venda)",
            tela, exp, "usd-brl",
        ))

    # 3. Dollar Index (DXY)
    s = macro["dxy"]
    if s:
        atual = s[-1]; pico = max(s, key=lambda r: r[1]); minima = min(s, key=lambda r: r[1])
        tela, exp = _chart_par(s, "Dólar Index — DXY (ICE)", "#3A6EA5", uid="dxy")
        blocos.append(chart_card(
            "Dólar Index (DXY)",
            f"Atual: <strong>{atual[1]:.2f}</strong> em {fmt_date_br(atual[0].isoformat())} · "
            f"Pico: <strong>{pico[1]:.2f}</strong> em {fmt_date_br(pico[0].isoformat())} · "
            f"Mínima: <strong>{minima[1]:.2f}</strong> · "
            f"Fonte: Yahoo Finance · DX-Y.NYB",
            tela, exp, "dollar-index",
        ))

    # 4. Balança Comercial Brasil (acum. 12m, US$ mi)
    s_mes = macro["balanca_br"]
    if s_mes:
        s = _balanca_acum12m(s_mes)
        atual = s[-1] if s else s_mes[-1]
        pico = max(s, key=lambda r: r[1]); minima = min(s, key=lambda r: r[1])
        tela, exp = _chart_par(
            s, "Balança Comercial BR · saldo acumulado 12 m (US$ milhões)",
            "#1E5631", uid="balbr", eixo_zero=True,
        )
        blocos.append(chart_card(
            "Balança Comercial Brasil (12 m)",
            f"Saldo 12 m: <strong>{_fmt_us_mi(atual[1])}</strong> em {fmt_date_br(atual[0].isoformat())} · "
            f"Pico: <strong>{_fmt_us_mi(pico[1])}</strong> · "
            f"Mínima: <strong>{_fmt_us_mi(minima[1])}</strong> · "
            f"Fonte: BCB SGS · série 22707",
            tela, exp, "balanca-br",
        ))

    # 5. Balança Comercial EUA (anual, US$ tri)
    s = macro["balanca_us"]
    if s:
        # Convert para US$ bi pra visualização (valores são em US$)
        s_bi = [(d, v / 1e9) for d, v in s]
        atual = s_bi[-1]
        pico = max(s_bi, key=lambda r: r[1]); minima = min(s_bi, key=lambda r: r[1])
        tela, exp = _chart_par(
            s_bi, "Balança Comercial EUA · anual (US$ bilhões)",
            "#8B2D3C", uid="balus", eixo_zero=True,
        )
        blocos.append(chart_card(
            "Balança Comercial EUA (anual)",
            f"Último ({atual[0].year}): <strong>{_fmt_us_bi(atual[1] * 1e9, 1)}</strong> · "
            f"Pico: <strong>{_fmt_us_bi(pico[1] * 1e9, 1)}</strong> ({pico[0].year}) · "
            f"Mínima: <strong>{_fmt_us_bi(minima[1] * 1e9, 1)}</strong> ({minima[0].year}) · "
            f"Fonte: World Bank · NE.RSB.GNFS.CD",
            tela, exp, "balanca-us",
        ))

    # 6. Fluxo de Capital Estrangeiro (portfolio · mensal · US$ mi)
    s = macro["fluxo_cap"]
    if s:
        atual = s[-1]; pico = max(s, key=lambda r: r[1]); minima = min(s, key=lambda r: r[1])
        tela, exp = _chart_par(
            s, "Investimento Estrangeiro Portfolio · líq. mensal (US$ mi)",
            "#6B46C1", uid="fluxo", eixo_zero=True,
        )
        blocos.append(chart_card(
            "Fluxo de Capital Estrangeiro (portfolio)",
            f"Último: <strong>{_fmt_us_mi(atual[1])}</strong> em {fmt_date_br(atual[0].isoformat())} · "
            f"Pico de entrada: <strong>{_fmt_us_mi(pico[1])}</strong> · "
            f"Pico de saída: <strong>{_fmt_us_mi(minima[1])}</strong> · "
            f"Fonte: BCB SGS · série 11759",
            tela, exp, "fluxo-cap",
        ))

    return "\n".join(blocos)


def html_focus(macro: dict) -> str:
    """Tabela compacta com as expectativas Focus para os principais indicadores."""
    focus = macro.get("focus", {})
    if not focus:
        return ""

    indicadores = [
        ("Selic",      "Selic-Meta (% a.a.)",  "%"),
        ("IPCA",       "IPCA (% acum. ano)",   "%"),
        ("PIB Total",  "PIB Total (% a.a.)",   "%"),
        ("Câmbio",     "USD/BRL (fim período)", "R$"),
    ]
    anos = sorted({a for ind in focus.values() for a in ind.keys()})[:5]

    head = "<tr><th class=\"col-comite\">Indicador</th>"
    for a in anos:
        head += f"<th class=\"col-proj\">{a}</th>"
    head += "<th class=\"col-faltam\">Data</th></tr>"

    rows: list[str] = []
    data_ref = ""
    for ind_key, ind_label, unidade in indicadores:
        if ind_key not in focus:
            continue
        cells = [f"<td class=\"col-comite\">{ind_label}</td>"]
        for a in anos:
            r = focus[ind_key].get(a)
            if r and r.get("mediana") is not None:
                v = r["mediana"]
                if unidade == "%":
                    s = fmt_pct_br(v, 2)
                else:
                    s = f"R$ {v:.2f}".replace(".", ",")
                cells.append(f"<td class=\"col-proj\">{s}</td>")
                data_ref = r.get("data") or data_ref
            else:
                cells.append("<td class=\"col-proj\">—</td>")
        # Data da leitura na última coluna
        cells.append(f"<td class=\"col-faltam\">{fmt_date_br(data_ref) if data_ref else '—'}</td>")
        rows.append("<tr>" + "".join(cells) + "</tr>")

    return (
        '<div class="tabela-wrap" data-tabela="focus">'
        '<table><thead>'
        + head
        + '</thead><tbody>'
        + "".join(rows)
        + '</tbody></table></div>'
    )


# ---------------------------------------------------------------------------
# HTML completo
# ---------------------------------------------------------------------------

def render_html() -> str:
    fed_atual = FOMC[-1][1]
    selic_atual = SELIC[-1][1]
    diff_atual = selic_atual - fed_atual

    fed_max = max(FOMC, key=lambda r: r[1])
    fed_min = min(FOMC, key=lambda r: r[1])
    selic_max = max(SELIC, key=lambda r: r[1])
    selic_min = min(SELIC, key=lambda r: r[1])

    # Séries para charts
    serie_fed = [(parse_date(d), v) for d, v, *_ in FOMC]
    serie_selic = [(parse_date(d), v) for d, v, *_ in SELIC]

    inicio = dt.date(2003, 1, 1)
    fim = HOJE
    s_fed_m = serie_mensal(FOMC, inicio, fim)
    s_sel_m = serie_mensal(SELIC, inicio, fim)
    serie_diff = [(d, sv - fv) for (d, sv), (_, fv) in zip(s_sel_m, s_fed_m)]

    chart_fed = chart_linha(
        serie_fed,
        "Federal Funds Rate (FOMC)",
        "#3A6EA5",
        uid="fed",
    )
    chart_sel = chart_linha(
        serie_selic,
        "Selic-Meta (Copom · BCB)",
        "#1E5631",
        uid="sel",
    )
    chart_dif = chart_linha(
        serie_diff,
        "Diferencial Selic – Fed Funds (p.p.)",
        "#8B2D3C",
        uid="dif",
        eixo_zero=True,
    )
    # Versões dedicadas para o botão "copiar imagem" — proporção mais
    # quadrada e fontes maiores, recalculadas para não sobrepor.
    chart_fed_x = chart_linha(serie_fed,
        "Federal Funds Rate (FOMC)", "#3A6EA5", uid="fedx", export=True)
    chart_sel_x = chart_linha(serie_selic,
        "Selic-Meta (Copom · BCB)", "#1E5631", uid="selx", export=True)
    chart_dif_x = chart_linha(serie_diff,
        "Diferencial Selic – Fed Funds (p.p.)", "#8B2D3C", uid="difx",
        eixo_zero=True, export=True)

    # ---- Carrega séries macro (Dívida, câmbio, balanças, fluxo, Focus)
    macro = _carregar_macro()
    macro_html = html_secao_macro(macro)
    focus_html = html_focus(macro)

    kpis = f"""
    <div class="kpi-grid">
      <div class="kpi">
        <div class="label">Fed Funds (atual)</div>
        <div class="value">{fmt_pct_br(fed_atual, 2)}</div>
        <div class="delta">Última decisão · {fmt_date_br(FOMC[-1][0])}</div>
      </div>
      <div class="kpi">
        <div class="label">Taxa Selic (a.a.)</div>
        <div class="value">{fmt_pct_br(selic_atual, 2)}</div>
        <div class="delta">Última decisão Copom · {fmt_date_br(SELIC[-1][0])}</div>
      </div>
      <div class="kpi">
        <div class="label">Diferencial (Selic – Fed)</div>
        <div class="value">{fmt_pct_br(diff_atual, 2).replace('%','')} p.p.</div>
        <div class="delta">Spread vigente</div>
      </div>
      <div class="kpi">
        <div class="label">Janela analisada</div>
        <div class="value">2003 → 2026</div>
        <div class="delta">{len(FOMC)} decisões FOMC · {len(SELIC)} Copom</div>
      </div>
    </div>"""

    proxima_fomc = next((d for d, *_ in PROXIMAS_FOMC if parse_date(d) >= HOJE), None)
    proxima_copom = next((d for d, *_ in PROXIMAS_COPOM if parse_date(d) >= HOJE), None)
    proxima_str = (
        f"FOMC: {fmt_date_br(proxima_fomc)} · Copom: {fmt_date_br(proxima_copom)}"
        if proxima_fomc and proxima_copom else "—"
    )

    css = """
:root {
  --navy: #0A2342;
  --green: #1E5631;
  --green-soft: #2E7D4F;
  --ink: #1B1F23;
  --muted: #5B6775;
  --line: #E4E7EB;
  --zebra: #F7F9FB;
  --accent: #F3F6F4;
  --warn: #B45309;
  --danger: #B91C1C;
  --up: #047857;
  --down: #B91C1C;
  --new: #FFFBEA;
  --new-br: #F59E0B;
  --alert-bg: #FEF3C7;
}
* { box-sizing: border-box; }
html, body {
  margin: 0; padding: 0; background: #FFFFFF; color: var(--ink);
  font-family: "Segoe UI", "Inter", "Helvetica Neue", Arial, sans-serif;
  font-size: 14px; line-height: 1.55;
  -webkit-font-smoothing: antialiased;
}
.page {
  max-width: 1720px; width: 100%; margin: 0 auto;
  padding: 40px clamp(24px, 3.5vw, 56px) 80px;
}

header.top {
  display: flex; align-items: center; justify-content: space-between;
  border-bottom: 3px solid var(--green); padding-bottom: 20px; margin-bottom: 22px;
  gap: 24px;
}
.brand { display: flex; align-items: center; gap: 18px; }
.brand .logo {
  height: 56px; width: auto; display: block;
  object-fit: contain; flex-shrink: 0;
}
.brand .divider { width: 1px; height: 56px; background: var(--line); }
.brand .tagline {
  font-size: 12px; color: var(--muted);
  letter-spacing: .12em; text-transform: uppercase;
}
.brand .tagline strong {
  color: var(--navy); font-size: 13px;
  letter-spacing: .14em; display: block;
}
.doc-meta { text-align: right; font-size: 12px; color: var(--muted); }
.doc-meta .date { color: var(--navy); font-weight: 700; font-size: 14px; }
.doc-meta .badge {
  display: inline-block; padding: 2px 8px; background: var(--green);
  color: #fff; border-radius: 3px; font-size: 10px; font-weight: 700;
  letter-spacing: .08em; text-transform: uppercase; margin-top: 4px;
}

h1.title { font-size: 30px; font-weight: 700; letter-spacing: -.01em;
  color: var(--navy); margin: 0 0 4px; }
.subtitle { font-size: 15px; font-weight: 600; color: var(--green);
  margin: 0 0 22px; }

h2 {
  font-size: 20px; background: var(--navy); color: #fff;
  padding: 12px 18px; border-left: 6px solid var(--green);
  letter-spacing: .02em; font-weight: 600; margin: 36px 0 18px;
}
h2 .sub { font-weight: 400; color: #C9D4E0; font-size: 13px; margin-left: 10px; }

.kpi-grid {
  display: grid; grid-template-columns: repeat(4, 1fr); gap: 14px; margin: 18px 0 22px;
}
@media (max-width: 960px) { .kpi-grid { grid-template-columns: repeat(2, 1fr); } }
.kpi {
  padding: 14px 16px; background: var(--accent);
  border-top: 3px solid var(--green); border-bottom: 1px solid var(--line);
}
.kpi .label { font-size: 11px; color: var(--muted);
  text-transform: uppercase; letter-spacing: .08em; }
.kpi .value { font-size: 22px; color: var(--navy); font-weight: 800; margin: 4px 0 2px; }
.kpi .delta { font-size: 12px; color: var(--green); font-weight: 600; }

.tabela-wrap {
  border: 1px solid var(--line); background: #fff; max-height: 460px;
  overflow-y: auto; box-shadow: 0 1px 2px rgba(10,35,66,.04);
}
table { width: 100%; border-collapse: collapse; font-size: 13px; background: #fff; }
thead th {
  position: sticky; top: 0; background: var(--green); color: #fff;
  font-weight: 600; padding: 10px 12px;
  font-size: 12px; letter-spacing: .03em; text-transform: uppercase;
  border-bottom: 2px solid var(--navy);
}
tbody td { padding: 9px 12px; border-bottom: 1px solid var(--line); }
tbody tr:nth-child(even) { background: var(--zebra); }
tbody tr:hover { background: #EEF4EF; }

/* Alinhamento por coluna semântica — header e células casados */
.col-data    { text-align: left;   font-weight: 600; color: var(--navy);
               white-space: nowrap; }
.col-taxa    { text-align: center; }
.col-mov     { text-align: left;   color: var(--muted); }
.col-comite  { text-align: left;   font-weight: 600; color: var(--navy); }
.col-proj    { text-align: left;   color: var(--ink); white-space: nowrap; }
.col-faltam  { text-align: right;  color: var(--muted); white-space: nowrap; }
thead th.col-data, thead th.col-comite { color: #fff; font-weight: 600; }

/* Duas tabelas lado a lado */
.duas-tabelas {
  display: grid; grid-template-columns: 1fr 1fr; gap: 22px; margin-top: 4px;
}
@media (max-width: 1100px) { .duas-tabelas { grid-template-columns: 1fr; } }
.tab-block { background: #fff; }
.tab-block-head {
  font-size: 13px; color: #fff;
  background: var(--navy);
  padding: 11px 16px;
  border-left: 4px solid var(--green);
  letter-spacing: .04em; font-weight: 600;
}
.tab-block-head strong {
  font-weight: 800; letter-spacing: .1em; text-transform: uppercase;
  color: #fff; margin-right: 6px;
}
.tab-block-head .sep { color: #C9D4E0; margin: 0 6px; }
.tab-block .tabela-wrap { border-top: none; }

.pill {
  display: inline-block; padding: 2px 9px; border-radius: 3px;
  font-size: 11px; font-weight: 700; letter-spacing: .04em;
  font-variant-numeric: tabular-nums;
}
.pill.alta   { background: #FEE2E2; color: var(--down); }
.pill.baixa  { background: #D1FAE5; color: var(--up); }
.pill.neutro { background: var(--line); color: var(--muted); }

.chart-card {
  border: 1px solid var(--line); background: #fff;
  padding: 14px 16px 8px; margin-bottom: 18px;
  box-shadow: 0 1px 2px rgba(10,35,66,.04);
}
.chart-card .chart-head {
  display: flex; align-items: flex-start; justify-content: space-between;
  gap: 16px; margin-bottom: 6px;
}
.chart-card .ch-title {
  font-size: 13px; font-weight: 700; color: var(--navy);
  letter-spacing: .06em; text-transform: uppercase; margin-bottom: 4px;
}
.chart-card .ch-meta { font-size: 11px; color: var(--muted); }
.chart-svg { width: 100%; height: auto; display: block; }

.copy-btn {
  display: inline-flex; align-items: center; justify-content: center;
  width: 28px; height: 28px; padding: 0;
  border: 1px solid var(--line); background: transparent;
  color: var(--muted); border-radius: 4px; cursor: pointer;
  font-family: inherit; flex-shrink: 0; opacity: .55;
  transition: color .15s, border-color .15s, background .15s, opacity .15s;
}
.copy-btn:hover  { color: var(--navy); border-color: var(--navy);
                   background: #fff; opacity: 1; }
.copy-btn:focus-visible { outline: 2px solid var(--navy); outline-offset: 1px; }
.copy-btn.ok    { color: var(--up);   border-color: var(--up);   opacity: 1; }
.copy-btn.fail  { color: var(--down); border-color: var(--down); opacity: 1; }
.copy-btn svg   { display: block; }
.chart-card:hover .copy-btn { opacity: 1; }
@media print { .copy-btn { display: none; } }

.alerts-box {
  border: 2px solid var(--new-br); background: var(--alert-bg);
  padding: 14px 18px; margin: 18px 0; font-size: 12.5px;
}
.alerts-box strong { color: var(--navy); }

footer {
  margin-top: 50px; padding-top: 16px; border-top: 1px solid var(--line);
  font-size: 11px; color: var(--muted);
  display: flex; justify-content: space-between; flex-wrap: wrap; gap: 12px;
}
footer .src strong { color: var(--navy); }

.two-col { display: grid; grid-template-columns: 1fr 1fr; gap: 18px; }
@media (max-width: 1024px) { .two-col { grid-template-columns: 1fr; } }

@media print {
  * { -webkit-print-color-adjust: exact !important;
      print-color-adjust: exact !important; }
  @page { margin: 1.2cm 1cm; }
  .tabela-wrap { max-height: none; overflow: visible; }
}
"""

    html = f"""<!doctype html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Juros — FOMC vs Selic</title>
<style>{css}</style>
</head>
<body>
<div class="page">
  <header class="top">
    <div class="brand">
      <div class="tagline">
        <strong>Market Intelligence</strong>
        Política monetária comparada · FOMC e Copom
      </div>
    </div>
    <div class="doc-meta">
      <div class="date">{HOJE.strftime('%d/%m/%Y')}</div>
      <div>Próximas: {proxima_str}</div>
      <div><span class="badge">● LIVE</span></div>
    </div>
  </header>

  <h1 class="title">Decisões de juros — Estados Unidos &amp; Brasil</h1>
  <div class="subtitle">FOMC (Federal Reserve) · Copom (Banco Central do Brasil) · 2003 → 2026</div>

  {kpis}

  <div class="alerts-box">
    <strong>Como ler este relatório.</strong> A série da Selic-Meta é puxada
    em tempo real do <strong>SGS/BCB · série 432</strong>; as datas mostradas
    representam o início de vigência da meta (em geral D+1 da reunião do Copom).
    Para o FOMC adotamos o ponto médio da banda-alvo do Fed Funds quando
    aplicável, com a data da decisão. As próximas reuniões seguem o calendário
    oficial. Os gráficos rotulam os pontos extremos (picos, vales, atual).
    Em caso de indisponibilidade da API do BCB, o relatório usa o último
    cache local e exibe um aviso no console do build.
  </div>

  <h2>Próximas divulgações <span class="sub">FOMC + Copom · ordem cronológica</span></h2>
  {html_tabela_proximas()}

  <h2>Decisões dos comitês <span class="sub">histórico · taxas-meta vigentes</span></h2>
  <div class="duas-tabelas">
    <div class="tab-block">
      <div class="tab-block-head">
        <strong>FOMC</strong> · Federal Reserve · Fed Funds (ponto médio do range)
      </div>
      {html_tabela_decisoes(FOMC, "fomc")}
    </div>
    <div class="tab-block">
      <div class="tab-block-head">
        <strong>Copom</strong> · Banco Central do Brasil · Selic-Meta
      </div>
      {html_tabela_decisoes(SELIC, "copom")}
    </div>
  </div>

  <h2>Evolução · Federal Funds Rate <span class="sub">2003 → hoje · pontos rotulados nos extremos</span></h2>
  {chart_card(
      "Fed Funds Rate · taxa-meta (FOMC)",
      f"Pico: <strong>{fmt_pct_br(fed_max[1])}</strong> em {fmt_date_br(fed_max[0])} · "
      f"Mínima: <strong>{fmt_pct_br(fed_min[1])}</strong> em {fmt_date_br(fed_min[0])} · "
      f"Atual: <strong>{fmt_pct_br(fed_atual)}</strong>",
      chart_fed,
      chart_fed_x,
      "fed-funds-rate",
  )}

  <h2>Evolução · Selic-Meta <span class="sub">2003 → hoje · Banco Central do Brasil</span></h2>
  {chart_card(
      "Selic-Meta · taxa-meta (Copom)",
      f"Pico: <strong>{fmt_pct_br(selic_max[1])}</strong> em {fmt_date_br(selic_max[0])} · "
      f"Mínima: <strong>{fmt_pct_br(selic_min[1])}</strong> em {fmt_date_br(selic_min[0])} · "
      f"Atual: <strong>{fmt_pct_br(selic_atual)}</strong>",
      chart_sel,
      chart_sel_x,
      "selic-meta",
  )}

  <h2>Diferencial Selic – Fed Funds <span class="sub">spread em pontos percentuais · forward-fill mensal</span></h2>
  {chart_card(
      "Spread Selic-Meta vs Fed Funds (p.p.)",
      f"Spread atual: <strong>{fmt_pct_br(diff_atual, 2).replace('%','')} p.p.</strong> · "
      f"Janela: 2003-01 → {HOJE.strftime('%Y-%m')}",
      chart_dif,
      chart_dif_x,
      "spread-selic-fed",
  )}

  <h2>Macro Brasil <span class="sub">dívida fiscal · câmbio · balança · fluxo de capital</span></h2>
  {macro_html}

  <h2>Expectativas Focus <span class="sub">mediana das projeções de mercado · BCB Olinda</span></h2>
  {focus_html}

  <footer>
    <div class="src">
      <strong>Fontes:</strong> Federal Reserve (FOMC statements / projections) ·
      Banco Central do Brasil (Atas e Comunicados Copom · SGS · Olinda Expectativas) ·
      Yahoo Finance (DX-Y.NYB) · World Bank Open Data (NE.RSB.GNFS.CD) ·
      compilação própria.
    </div>
    <div>© 2026 · Política monetária comparada · Não constitui recomendação de investimento.</div>
  </footer>
</div>

<script>
(function () {{
  // Layout do export — leve (~40–80 KB) e proporção ~16:10
  const EXPORT_W = 900;
  const HEADER_H = 110;
  const CHART_SLOT_H = 460;
  const FOOTER_H = 30;
  const PAD = 28;
  const SCALE = 1.0;       // pixels finais = EXPORT_W × totalH (sem upscale)

  function escapeXml(s) {{
    return String(s)
      .replace(/&/g, "&amp;").replace(/</g, "&lt;").replace(/>/g, "&gt;")
      .replace(/"/g, "&quot;").replace(/'/g, "&apos;");
  }}

  function wrapText(s, maxChars, maxLines) {{
    const words = s.split(/\\s+/);
    const lines = [];
    let cur = "";
    for (const w of words) {{
      const test = cur ? cur + " " + w : w;
      if (test.length > maxChars && cur) {{
        lines.push(cur);
        cur = w;
        if (lines.length >= maxLines - 1) break;
      }} else {{
        cur = test;
      }}
    }}
    if (cur && lines.length < maxLines) lines.push(cur);
    return lines;
  }}

  function buildExportSvg(card) {{
    // Usa o SVG dedicado (proporção mais quadrada, fontes maiores e
    // de-clutter recalculado), escondido em .ch-svg-export-wrap. A versão
    // visível na tela continua intocada.
    const orig = card.querySelector(
      ".ch-svg-export-wrap svg.chart-svg-export"
    ) || card.querySelector("svg.chart-svg");

    const titleEl = card.querySelector(".ch-title");
    const metaEl = card.querySelector(".ch-meta");
    const title = (titleEl ? titleEl.textContent : "").trim();
    const meta = (metaEl ? metaEl.textContent : "").replace(/\\s+/g, " ").trim();

    const innerVB = orig.getAttribute("viewBox") ||
      `0 0 ${{orig.viewBox.baseVal.width}} ${{orig.viewBox.baseVal.height}}`;

    const W = EXPORT_W;
    const totalH = HEADER_H + CHART_SLOT_H + FOOTER_H;
    const chartW = W - PAD * 2;
    const FONT = '"Segoe UI", "Inter", "Helvetica Neue", Arial, sans-serif';

    const metaLines = wrapText(meta, 78, 2);
    const innerXML = orig.innerHTML;

    const xml = `<?xml version="1.0" encoding="UTF-8"?>
<svg xmlns="http://www.w3.org/2000/svg" width="${{W}}" height="${{totalH}}"
     viewBox="0 0 ${{W}} ${{totalH}}">
  <rect width="${{W}}" height="${{totalH}}" fill="#FFFFFF"/>
  <rect x="0" y="0" width="6" height="${{totalH}}" fill="#1E5631"/>

  <text x="${{PAD}}" y="46" font-family='${{FONT}}' font-size="22" font-weight="700"
        fill="#0A2342" letter-spacing="0.5">${{escapeXml(title)}}</text>
  ${{metaLines.map((line, i) =>
    `<text x="${{PAD}}" y="${{74 + i*18}}" font-family='${{FONT}}' font-size="12.5"
       fill="#5B6775">${{escapeXml(line)}}</text>`
  ).join("")}}
  <line x1="${{PAD}}" y1="${{HEADER_H - 6}}" x2="${{W - PAD}}" y2="${{HEADER_H - 6}}"
        stroke="#E4E7EB" stroke-width="1"/>

  <svg x="${{PAD}}" y="${{HEADER_H + 8}}"
       width="${{chartW}}" height="${{CHART_SLOT_H - 16}}"
       viewBox="${{innerVB}}" preserveAspectRatio="xMidYMid meet">
    ${{innerXML}}
  </svg>

  <text x="${{W/2}}" y="${{totalH - 11}}" font-family='${{FONT}}' font-size="10"
        fill="#5B6775" text-anchor="middle" letter-spacing="0.18em"
        font-weight="600">MARKET INTELLIGENCE</text>
</svg>`;
    return {{ xml, w: W, h: totalH }};
  }}

  function svgToPngBlob(card) {{
    return new Promise((resolve, reject) => {{
      const {{ xml, w, h }} = buildExportSvg(card);
      const svgBlob = new Blob([xml], {{ type: "image/svg+xml;charset=utf-8" }});
      const url = URL.createObjectURL(svgBlob);
      const img = new Image();
      img.onload = () => {{
        const canvas = document.createElement("canvas");
        canvas.width = Math.round(w * SCALE);
        canvas.height = Math.round(h * SCALE);
        const ctx = canvas.getContext("2d");
        ctx.fillStyle = "#FFFFFF";
        ctx.fillRect(0, 0, canvas.width, canvas.height);
        ctx.drawImage(img, 0, 0, canvas.width, canvas.height);
        URL.revokeObjectURL(url);
        canvas.toBlob(b => b ? resolve(b) : reject(new Error("toBlob falhou")), "image/png");
      }};
      img.onerror = (e) => {{ URL.revokeObjectURL(url); reject(e); }};
      img.src = url;
    }});
  }}

  function downloadFallback(blob, filename) {{
    const url = URL.createObjectURL(blob);
    const a = document.createElement("a");
    a.href = url;
    a.download = filename + ".png";
    document.body.appendChild(a);
    a.click();
    a.remove();
    setTimeout(() => URL.revokeObjectURL(url), 1000);
  }}

  const ICON_OK = '<svg width="15" height="15" viewBox="0 0 24 24" fill="none"'
    + ' stroke="currentColor" stroke-width="2.5" stroke-linecap="round"'
    + ' stroke-linejoin="round"><polyline points="20 6 9 17 4 12"/></svg>';
  const ICON_FAIL = '<svg width="15" height="15" viewBox="0 0 24 24" fill="none"'
    + ' stroke="currentColor" stroke-width="2.5" stroke-linecap="round"'
    + ' stroke-linejoin="round"><line x1="18" y1="6" x2="6" y2="18"/>'
    + '<line x1="6" y1="6" x2="18" y2="18"/></svg>';

  async function handleCopy(btn) {{
    const card = btn.closest(".chart-card");
    if (!card || !card.querySelector("svg.chart-svg")) return;
    const filename = btn.dataset.filename || "juros-grafico";
    const originalIcon = btn.innerHTML;
    const originalTitle = btn.title;
    btn.disabled = true;
    btn.title = "Gerando…";

    try {{
      const blob = await svgToPngBlob(card);
      let copied = false;
      if (navigator.clipboard && window.ClipboardItem) {{
        try {{
          await navigator.clipboard.write([
            new ClipboardItem({{ "image/png": blob }})
          ]);
          copied = true;
        }} catch (_) {{ copied = false; }}
      }}
      if (copied) {{
        btn.classList.add("ok");
        btn.innerHTML = ICON_OK;
        btn.title = "Copiado!";
      }} else {{
        downloadFallback(blob, filename);
        btn.classList.add("ok");
        btn.innerHTML = ICON_OK;
        btn.title = "Baixado (PNG)";
      }}
    }} catch (err) {{
      console.error("Falha ao gerar PNG:", err);
      btn.classList.add("fail");
      btn.innerHTML = ICON_FAIL;
      btn.title = "Falhou — tente novamente";
    }} finally {{
      setTimeout(() => {{
        btn.classList.remove("ok", "fail");
        btn.innerHTML = originalIcon;
        btn.title = originalTitle;
        btn.disabled = false;
      }}, 2000);
    }}
  }}

  document.querySelectorAll(".copy-btn").forEach(b => {{
    b.addEventListener("click", () => handleCopy(b));
  }});
}})();
</script>
</body>
</html>
"""
    return html


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

NAVY = "FF0A2342"
GREEN = "FF1E5631"
WHITE = "FFFFFFFF"
ZEBRA = "FFF7F9FB"
LINE = "FFE4E7EB"
ACCENT = "FFF3F6F4"


def _style_header(ws, ncols: int):
    fill = PatternFill("solid", fgColor=GREEN)
    side_navy = Side(style="medium", color=NAVY)
    side_line = Side(style="thin", color=LINE)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = Font(name="Segoe UI", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=side_navy, left=side_line, right=side_line)
    ws.row_dimensions[1].height = 22


def _zebra(ws, nrows: int, ncols: int):
    side = Side(style="thin", color=LINE)
    for r in range(2, nrows + 2):
        bg = ZEBRA if r % 2 == 0 else None
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name="Segoe UI", size=10, color="FF1B1F23")
            cell.border = Border(bottom=side, left=side, right=side)
            if bg:
                cell.fill = PatternFill("solid", fgColor=bg)
            if c == 1:
                cell.font = Font(name="Segoe UI", size=10, bold=True, color=NAVY)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")


def _write_decisoes(ws, headers: list[str], data: list[tuple]):
    ws.append(headers)
    for row in data:
        d_iso, taxa, _evento, obs = row
        ws.append([
            dt.date.fromisoformat(d_iso),
            taxa / 100.0,    # número como fração para usar formato % do Excel
            obs,
        ])
    _style_header(ws, len(headers))
    _zebra(ws, len(data), len(headers))

    for r in range(2, len(data) + 2):
        ws.cell(row=r, column=1).number_format = "DD/MM/YYYY"
        ws.cell(row=r, column=2).number_format = "0.00%"

    widths = [18, 14, 56]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _write_proximas(ws):
    headers = ["Data prevista", "Comitê", "Projeção", "Faltam (dias)"]
    pares = []
    for d, _evt, proj in PROXIMAS_FOMC:
        pares.append((dt.date.fromisoformat(d), "FOMC · Federal Reserve", proj))
    for d, _evt, proj in PROXIMAS_COPOM:
        pares.append((dt.date.fromisoformat(d), "Copom · BCB", proj))
    pares.sort(key=lambda p: p[0])

    ws.append(headers)
    for d, comite, proj in pares:
        dias = (d - HOJE).days
        ws.append([d, comite, proj or "—", dias])

    _style_header(ws, len(headers))
    _zebra(ws, len(pares), len(headers))
    for r in range(2, len(pares) + 2):
        ws.cell(row=r, column=1).number_format = "DD/MM/YYYY"
        ws.cell(row=r, column=4).number_format = "0"
    widths = [16, 32, 32, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _write_diferencial(ws):
    headers = ["Mês", "Fed Funds (a.a.)", "Selic-Meta (a.a.)", "Diferencial (p.p.)"]
    inicio = dt.date(2003, 1, 1)
    fim = HOJE
    fed_m = serie_mensal(FOMC, inicio, fim)
    sel_m = serie_mensal(SELIC, inicio, fim)
    rows = [(d, f, s, s - f) for (d, f), (_, s) in zip(fed_m, sel_m)]

    ws.append(headers)
    for d, f, s, diff in rows:
        ws.append([d, f / 100.0, s / 100.0, diff / 100.0])

    _style_header(ws, len(headers))
    _zebra(ws, len(rows), len(headers))
    for r in range(2, len(rows) + 2):
        ws.cell(row=r, column=1).number_format = "MMM/YYYY"
        for c in (2, 3, 4):
            ws.cell(row=r, column=c).number_format = "0.00%"
    widths = [14, 18, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _write_resumo(ws):
    headers = ["Indicador", "Valor", "Referência"]
    fed_atual = FOMC[-1]
    selic_atual = SELIC[-1]
    fed_max = max(FOMC, key=lambda r: r[1])
    fed_min = min(FOMC, key=lambda r: r[1])
    selic_max = max(SELIC, key=lambda r: r[1])
    selic_min = min(SELIC, key=lambda r: r[1])

    diff = (selic_atual[1] - fed_atual[1]) / 100.0

    rows = [
        ("Fed Funds — taxa atual", fed_atual[1] / 100.0, f"Decisão de {fmt_date_br(fed_atual[0])}"),
        ("Fed Funds — pico do período", fed_max[1] / 100.0, fmt_date_br(fed_max[0])),
        ("Fed Funds — mínima do período", fed_min[1] / 100.0, fmt_date_br(fed_min[0])),
        ("Selic — taxa atual", selic_atual[1] / 100.0, f"Decisão de {fmt_date_br(selic_atual[0])}"),
        ("Selic — pico do período", selic_max[1] / 100.0, fmt_date_br(selic_max[0])),
        ("Selic — mínima do período", selic_min[1] / 100.0, fmt_date_br(selic_min[0])),
        ("Diferencial atual (Selic – Fed)", diff, "Pontos percentuais"),
        ("Total de decisões FOMC mapeadas", len(FOMC), "Janela 2003-2026"),
        ("Total de decisões Copom mapeadas", len(SELIC), "Janela 2003-2026"),
        ("Próxima reunião FOMC", dt.date.fromisoformat(PROXIMAS_FOMC[0][0]), PROXIMAS_FOMC[0][1]),
        ("Próxima reunião Copom", dt.date.fromisoformat(PROXIMAS_COPOM[0][0]), PROXIMAS_COPOM[0][1]),
    ]

    ws.append(headers)
    for label, valor, ref in rows:
        ws.append([label, valor, ref])
    _style_header(ws, len(headers))
    _zebra(ws, len(rows), len(headers))

    for r, (_, valor, _) in enumerate(rows, start=2):
        cell = ws.cell(row=r, column=2)
        if isinstance(valor, dt.date):
            cell.number_format = "DD/MM/YYYY"
        elif isinstance(valor, int):
            cell.number_format = "0"
        else:
            cell.number_format = "0.00%" if abs(valor) < 1 else "0.00"

    widths = [40, 18, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _write_serie_macro(ws, header_valor: str, serie: list[tuple[str, float]],
                       fmt_valor: str = "0.0000",
                       largura_valor: int = 18) -> None:
    """Aba simples: Data + Valor. Aceita serie [(iso, float)]."""
    ws.append(["Data", header_valor])
    for d, v in serie:
        ws.append([dt.date.fromisoformat(d), v])
    _style_header(ws, 2)
    _zebra(ws, len(serie), 2)
    for r in range(2, len(serie) + 2):
        ws.cell(row=r, column=1).number_format = "DD/MM/YYYY"
        ws.cell(row=r, column=2).number_format = fmt_valor
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = largura_valor
    ws.freeze_panes = "A2"


def _write_focus(ws, focus: dict) -> None:
    """Aba Focus — pivota por indicador x ano de referência (mediana)."""
    if not focus:
        ws.append(["Focus indisponível"])
        return
    anos = sorted({a for ind in focus.values() for a in ind.keys()})[:6]
    ws.append(["Indicador", *anos, "Data da leitura"])
    for ind_label, dados_ano in focus.items():
        linha: list = [ind_label]
        data_ref = ""
        for a in anos:
            r = dados_ano.get(a, {})
            linha.append(r.get("mediana"))
            data_ref = r.get("data") or data_ref
        linha.append(dt.date.fromisoformat(data_ref) if data_ref else None)
        ws.append(linha)
    _style_header(ws, len(anos) + 2)
    _zebra(ws, len(focus), len(anos) + 2)
    last_col = len(anos) + 2
    for r in range(2, len(focus) + 2):
        ws.cell(row=r, column=last_col).number_format = "DD/MM/YYYY"
        for c in range(2, last_col):
            ws.cell(row=r, column=c).number_format = "0.00"
    ws.column_dimensions["A"].width = 22
    for i in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.freeze_panes = "B2"


def render_xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    _write_resumo(ws)

    _write_decisoes(
        wb.create_sheet("FOMC · Fed Funds"),
        ["Data da decisão", "Taxa-meta", "Movimento"],
        FOMC,
    )
    _write_decisoes(
        wb.create_sheet("Copom · Selic"),
        ["Data da decisão", "Taxa-meta", "Movimento"],
        SELIC,
    )
    _write_diferencial(wb.create_sheet("Diferencial mensal"))
    _write_proximas(wb.create_sheet("Próximas reuniões"))

    # ---- abas macro (não falham o build se a série estiver vazia)
    macro = _carregar_macro()
    if macro.get("divida_pib"):
        _write_serie_macro(
            wb.create_sheet("Dívida-PIB BR"),
            "Dívida Bruta / PIB (%)",
            [(d.isoformat(), v) for d, v in macro["divida_pib"]],
            fmt_valor="0.00",
        )
    if macro.get("usd_brl_d"):
        _write_serie_macro(
            wb.create_sheet("USD-BRL diária"),
            "PTAX venda (R$)",
            [(d.isoformat(), v) for d, v in macro["usd_brl_d"]],
            fmt_valor="0.0000",
        )
    if macro.get("dxy"):
        _write_serie_macro(
            wb.create_sheet("Dollar Index"),
            "DXY (pontos)",
            [(d.isoformat(), v) for d, v in macro["dxy"]],
            fmt_valor="0.00",
        )
    if macro.get("balanca_br"):
        _write_serie_macro(
            wb.create_sheet("Balança BR"),
            "Saldo mensal (US$ milhões)",
            [(d.isoformat(), v) for d, v in macro["balanca_br"]],
            fmt_valor="#,##0.0",
            largura_valor=22,
        )
    if macro.get("balanca_us"):
        _write_serie_macro(
            wb.create_sheet("Balança US"),
            "Saldo anual (US$)",
            [(d.isoformat(), v) for d, v in macro["balanca_us"]],
            fmt_valor="#,##0",
            largura_valor=24,
        )
    if macro.get("fluxo_cap"):
        _write_serie_macro(
            wb.create_sheet("Fluxo Capital BR"),
            "Inv. portfolio líq. (US$ milhões)",
            [(d.isoformat(), v) for d, v in macro["fluxo_cap"]],
            fmt_valor="#,##0.0",
            largura_valor=26,
        )
    if macro.get("focus"):
        _write_focus(wb.create_sheet("Focus"), macro["focus"])

    wb.save(path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    html = render_html()
    (ROOT / "index.html").write_text(html, encoding="utf-8")
    render_xlsx(ROOT / "juros.xlsx")
    print("OK")
    print(f"  - {ROOT / 'index.html'}")
    print(f"  - {ROOT / 'juros.xlsx'}")


if __name__ == "__main__":
    main()
